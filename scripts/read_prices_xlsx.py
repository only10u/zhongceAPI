"""一次性从 xlsx 导出站点与三模型价，供生成 seed_sites 用。依赖: pip install openpyxl"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

_DEFAULT_XLSX = Path(r"c:\Users\a2493\Desktop\中测 中转站价格大全（简洁）.xlsx")
XLSX = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else _DEFAULT_XLSX
OUT = Path(__file__).resolve().parents[1] / "data" / "_prices_export.json"


def _first_usd_block(s: object) -> str:
    """从单元格长文案里抓第一个 $ 金额（美金），失败则返回空。"""
    if s is None:
        return ""
    t = str(s)
    # 取如 $5.0000、$5.50
    m = re.search(r"\$\s*([\d.]+)", t.replace("\n", " "))
    if m:
        return m.group(1).rstrip("0").rstrip(".") or m.group(1)
    return ""


def _row_prices(row: tuple[object, ...]) -> tuple[str, str, str]:
    """Opus 4.7, Opus 4.6, Sonnet 4.6 三列的首个美金数字。"""
    cells = list(row)
    while len(cells) < 8:
        cells.append(None)
    o47 = _first_usd_block(cells[5])
    o46 = _first_usd_block(cells[6])
    s46 = _first_usd_block(cells[7])
    return (o47, o46, s46)


def compact_price(o47: str, o46: str, s46: str, max_len: int = 64) -> str:
    """压缩进 site_price VARCHAR(64)，示例：O4.7:$5 O4.6:$5 S4.6:$3"""
    parts = []
    if o47:
        parts.append(f"O4.7:${o47}")
    if o46:
        parts.append(f"O4.6:${o46}")
    if s46:
        parts.append(f"S4.6:${s46}")
    s = " ".join(parts)
    if len(s) <= max_len:
        return s
    # 再缩短：去 O4.7 前缀重复
    s2 = "/".join([p for p in (o47, o46, s46) if p])
    s2 = f"${s2}" if s2 else ""
    if len(s2) <= max_len:
        return s2
    return s[: max_len - 1] + "…"


def main() -> None:
    if not XLSX.is_file():
        raise SystemExit(f"找不到表格: {XLSX}")
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0] if rows else ()
    out_rows: list[dict[str, str | None]] = []

    # 合并逻辑：下一行若 name/url 空且三模型列有数，则并入上一行（DawCode 跨行）
    pending: dict | None = None
    for i, row in enumerate(rows[1:], start=2):
        name = row[0] if row[0] is not None else None
        bu = row[1] if len(row) > 1 else None
        name_s = str(name).strip() if name else ""
        bu_s = str(bu).strip() if bu else ""

        o47, o46, s46 = _row_prices(row)
        has_price = bool(o47 or o46 or s46)

        if name_s or bu_s:
            if pending:
                out_rows.append(pending)
            pending = {
                "sheet_row": i,
                "name": name_s or None,
                "base_url": bu_s or None,
                "opus47_usd": o47 or None,
                "opus46_usd": o46 or None,
                "sonnet46_usd": s46 or None,
            }
        elif pending and has_price:
            # 续行：补充价格（DawCode+ccmax）
            for k, v in (
                ("opus47_usd", o47),
                ("opus46_usd", o46),
                ("sonnet46_usd", s46),
            ):
                if v and not pending.get(k):
                    pending[k] = v
    if pending:
        out_rows.append(pending)

    # 生成 site_price 与 matched seed
    for r in out_rows:
        o47 = r.get("opus47_usd") or ""
        o46 = r.get("opus46_usd") or ""
        s46 = r.get("sonnet46_usd") or ""
        r["site_price"] = compact_price(str(o47), str(o46), str(s46))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        json.dumps({"header": [str(x) if x is not None else "" for x in header], "sites": out_rows}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(OUT.as_posix())


if __name__ == "__main__":
    main()
